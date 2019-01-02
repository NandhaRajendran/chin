from tkinter import *
import os, zipfile, shutil
from xml.dom import minidom

def getSourceFilePath() :
    strSourceFilePath = str(e1.get())
    return strSourceFilePath

def getArchiveFilePath() :
    strArchiveFilePath = str(e2.get())
    return strArchiveFilePath

def getResultFilePath() :
    strResultFilePath = str(e3.get())
    return strResultFilePath

def getFolderDate() :
    strFolderDate = str(e4.get())
    return strFolderDate

def setFileCount(intFoundFile) :
    v.set(intFoundFile)

def ReadExcelColumnAndToDictionary(path, strKeyColumn, strValueColumn):
    import openpyxl
    objExcelWorkBook = openpyxl.load_workbook(path)
    objExcelSheet = objExcelWorkBook.active
    objExcelValuesDictionary = {}
    intRowCount = objExcelSheet.max_row
    for i in range(2, intRowCount + 1):
        strKeyString = objExcelSheet.cell(row=i, column=strKeyColumn).value
        strValueString = objExcelSheet.cell(row=i, column=strValueColumn).value
        if not strKeyString in objExcelValuesDictionary:
            objExcelValuesDictionary[strKeyString]=[]
            objExcelValuesDictionary[strKeyString].append(strValueString)
    return objExcelValuesDictionary

def POSFileCheck():
    strCheckedStores = ""
    strSourcePath = getSourceFilePath()
    strArchivePath = getArchiveFilePath()
    strResultFilePath = getResultFilePath()
    strFolderString = getFolderDate()
    objSourceDictionary = ReadExcelColumnAndToDictionary(strSourcePath,3,2)
    #os.mkdir('.\\test')
    strResultFileCount = ''
    for key in objSourceDictionary :
        strSite = objSourceDictionary(key)
        if strSite in strCheckedStores :
        else:
            strCheckedStores = strCheckedStores & ";" & strSite
            strCountryName = strSite[0:3]
            strStoreID = strSite[-1:3]
            strZipFileDirectory = strArchivePath+'\\'+strCountryName+'\\' +strStoreID+ '\\' + strFolderString.split(':')[0] + '\\' + strFolderString.split(':')[1] + '\\' + strFolderString.split(':')[2]
            for filename in os.listdir(strZipFileDirectory):
                if filename.endswith(".zip"):
                    print(filename)
                    name = os.path.splitext(os.path.basename(filename))[0]
                    if not os.path.isdir(name):
                        zip = zipfile.ZipFile(filename)
                        # os.mkdir(name)
                        zip.extractall('C:\\Users\\tempnaraj\\Log\\Local')
            for xmlfiles in os.listdir('C:\\Users\\tempnaraj\\Log\\Local'):
                doc = minidom.parse(f'C:\\Users\\tempnaraj\\Log\\Local\\{xmlfiles}')
                # doc.getElementsByTagName returns NodeList
                name = doc.getElementsByTagName("title")[0]
                print(name.firstChild.data)
                if name.firstChild.data in objSourceDictionary:
                    shutil.copyfile(f'C:\\Users\\tempnaraj\\Log\\Local\\{xmlfiles}',f'C:\\Users\\tempnaraj\\Log\\Final\\{xmlfiles}')
            for files in os.listdir('C:\\Users\\tempnaraj\\Log\\Final'):
                shutil.copyfile(f'C:\\Users\\tempnaraj\\Log\\Final\\{files}', f'{strResultFilePath}\\{files}')
                strResultFileCount += 1

master = Tk()
master.title("POSLog Verification")
master.geometry('300x300')
Label(master, text="Source File Path").grid(row=0)
Label(master, text="Archive File Path").grid(row=2)
Label(master, text="Result File Path").grid(row=4)
Label(master, text="FolderDate(YY:MM:DD)").grid(row=6)
Label(master, text="Log Match Count").grid(row=8)
e1 = Entry(master)
e2 = Entry(master)
e3 = Entry(master)
e4 = Entry(master)
e5 = Entry(master)
v = StringVar()
e5 = Entry(master, textvariable=v)
e1.grid(row=0, column=2)
e2.grid(row=2, column=2)
e3.grid(row=4, column=2)
e4.grid(row=6, column=2)
e5.grid(row=8, column=2)
e6 = Button(master,text="Check it", command=POSFileCheck,bg='aqua')
e6.grid(row = 10,column=2)
mainloop()
