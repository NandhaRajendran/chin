# import openpyxl
#path = "C:\\Users\\Nandhakumar\\Documents\\Project\\file.xlsx"
# wb_obj = openpyxl.load_workbook(path) 
# sheet_obj = wb_obj.active 
# user_and_passs = {}
# intRowCount = sheet_obj.max_row
# for i in range(2, intRowCount + 1):
#     user = sheet_obj.cell(row=i, column=2).value
#     pass1 = sheet_obj.cell(row=i, column=3).value  
# #Problem seems to be in this general area
#     if not user in user_and_passs:
#         user_and_passs[user]=[]
#         user_and_passs[user].append(pass1)    
asi = input('name')
# Set Dictionary
def ReadExcelColumnAndToDictionary(path, strKeyColumn, strValueColumn):
    import openpyxl
    wb_obj = openpyxl.load_workbook(path) 
    sheet_obj = wb_obj.active 
    user_and_passs = {}
    intRowCount = sheet_obj.max_row
    for i in range(2, intRowCount + 1):
        user = sheet_obj.cell(row=i, column=strKeyColumn).value
        pass1 = sheet_obj.cell(row=i, column=strValueColumn).value  
    #Problem seems to be in this general area
        if not user in user_and_passs:
            user_and_passs[user]=[]
            user_and_passs[user].append(pass1)
    return user_and_passs