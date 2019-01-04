from tkinter import *
import os, zipfile, shutil
from xml.dom import expatbuilder as minidom

def ClearSourceFilePath() :
    e1.delete(0, END)

def ClearArchiveFilePath() :
    e2.delete(0, END)

def ClearResultFilePath() :
    e3.delete(0, END)

def ClearLogFolderPath() :
    e4.delete(0, END)

def ClearFolderDate() :
    e5.delete(0, END)

def ClearAllFilePath() :
    e1.delete(0, END)
    e2.delete(0, END)
    e3.delete(0, END)
    e4.delete(0, END)
    e5.delete(0, END)

def getSourceFilePath() :
    strSourceFilePath = str(e1.get())
    return strSourceFilePath

def getArchiveFilePath() :
    strArchiveFilePath = str(e2.get())
    return strArchiveFilePath

def getResultFilePath() :
    strResultFilePath = str(e3.get())
    return strResultFilePath

def getLogFilePath() :
    strLogFilePath = str(e4.get())
    return strLogFilePath

def getFolderDate() :
    strFolderDate = str(e5.get())
    return strFolderDate

def ReadExcelColumnAndToDictionary(path, strKeyColumn, strValueColumn):
    import openpyxl
    objExcelWorkBook = openpyxl.load_workbook(path)
    objExcelSheet = objExcelWorkBook.active
    objExcelValuesDictionary = {}
    intRowCount = objExcelSheet.max_row
    for i in range(2, intRowCount + 1):
        strKeyString = objExcelSheet.cell(row=i, column=strKeyColumn).value
        strValueString = objExcelSheet.cell(row=i, column=strValueColumn).value
        if not strKeyString in objExcelValuesDictionary:
            objExcelValuesDictionary[strKeyString]=strValueString
            #objExcelValuesDictionary[strKeyString].append(strValueString)
    return objExcelValuesDictionary

def POSFileCheck():
    strCheckedStores = ""
    strSourcePath = getSourceFilePath()
    strArchivePath = getArchiveFilePath()
    strResultFilePath = getResultFilePath()
    strFolderString = getFolderDate()
    strLogFilePath = getLogFilePath()
    #strAllZipFilesPath = f'{strLogFilePath}/Temporary'
    strAllXmlFilePath = f'{strLogFilePath}/Local'
    strFinalXmlPath = f'{strLogFilePath}/Final'
    if not strLogFilePath[:-1].endswith('/'):
        strLogFilePath = f'{strLogFilePath}/'
    #if not os.path.exists(strAllZipFilesPath):
     #   os.mkdir(strAllZipFilesPath)
    if not os.path.exists(strAllXmlFilePath):
        os.mkdir(strAllXmlFilePath)
    if not os.path.exists(strFinalXmlPath):
        os.mkdir(strFinalXmlPath)
    objSourceDictionary = {}
    objSourceDictionary = ReadExcelColumnAndToDictionary(strSourcePath,3,2)
    #os.mkdir('./test')
    for key in objSourceDictionary :
        strSite = objSourceDictionary[key]
        if strCheckedStores.find(str(strSite)) == -1:
            strCheckedStores = f'{strCheckedStores};{strSite}'
            #strCountryName = strSite[:2]
            #strStoreID = strSite[2:]
            #strZipFileDirectory = strArchivePath+'/'+strSite[:2]+'/' + strSite[2:] + '/' + strFolderString.split(':')[0] + '/' + strFolderString.split(':')[1] + '/' + strFolderString.split(':')[2]
            strZipFileDirectory = f"{strArchivePath}/{strSite[:2]}/{strSite[2:]}/{strFolderString.split(':')[0]}/{strFolderString.split(':')[1]}/{strFolderString.split(':')[2]}"
            for filename in os.listdir(strZipFileDirectory):
                if filename.endswith(".zip"):
                    print(filename)
                    name = os.path.splitext(os.path.basename(filename))[0]
                    if not os.path.isdir(name):
                        zip = zipfile.ZipFile(strZipFileDirectory+'/'+filename)
                        # os.mkdir(name)
                        zip.extractall(strAllXmlFilePath)
    for xmlfiles in os.listdir(strAllXmlFilePath):
        doc = minidom.parse(f'{strAllXmlFilePath}/{xmlfiles}', False)
        # doc.getElementsByTagName returns NodeList
        name = doc.getElementsByTagName("tns:UniqueTransactionID")[0]
        print(name.firstChild.data)
        if name.firstChild.data in objSourceDictionary:
            shutil.copyfile(f'{strAllXmlFilePath}/{xmlfiles}',f'{strFinalXmlPath}/{xmlfiles}')
    for files in os.listdir(strFinalXmlPath):
        shutil.copyfile(f'{strFinalXmlPath}/{files}', f'{strResultFilePath}/{files}')
    for Folder in os.listdir(strLogFilePath):
        shutil.rmtree(os.path.join(strLogFilePath, Folder))
master = Tk()
master.title("POSLog Verification")
master.geometry("1100x300+100+200")
master.grid_rowconfigure(0, weight=1)
master.grid_rowconfigure(2, weight=1)
master.grid_rowconfigure(4, weight=1)
master.grid_rowconfigure(6, weight=1)
master.grid_rowconfigure(8, weight=1)
master.grid_rowconfigure(10, weight=1)
master.grid_columnconfigure(0, minsize=150, weight=2)
master.grid_columnconfigure(2, minsize=600, weight=2)
master.grid_columnconfigure(4, minsize=600, weight=2)
master.grid_columnconfigure(6, minsize=600, weight=2)

l1 = Label(master, text="Source File Path",font=("Tahoma", 10, 'normal')).grid(row=0,sticky=E, pady=5)
l2 = Label(master, text="Archive File Path",font=("Tahoma", 10, 'normal')).grid(row=2,sticky=E, pady=5)
l3 = Label(master, text="Result File Path",font=("Tahoma", 10, 'normal')).grid(row=4,sticky=E, pady=5)
l4 = Label(master, text="Temporary Local Path",font=("Tahoma", 10, 'normal')).grid(row=6,sticky=E, pady=5)
l5 = Label(master, text="FolderDate(YY:MM:DD)",font=("Tahoma", 10, 'normal')).grid(row=8,sticky=E, pady=5)
l6 = Label(master).grid(row=10)
l7 = Label(master).grid(row=14)

e1 = Entry(master, width='150')
e2 = Entry(master, width='150')
e3 = Entry(master, width='150')
e4 = Entry(master, width='150')
e5 = Entry(master, width='50')


e1.grid(row=0, column=2,sticky=W)
e2.grid(row=2, column=2,sticky=W)
e3.grid(row=4, column=2,sticky=W)
e4.grid(row=6, column=2,sticky=W)
e5.grid(row=8, column=2,sticky=W)

e1.insert(END,"C:/Users/Nandhakumar/Documents/Project/file.xlsx")
e2.insert(END,"E:/Archive")
e3.insert(END,"E:/Obsolete")
e4.insert(END,"C:/Users/tempnaraj/Log")
e5.insert(END,"18:10:21")

c7 = Button(master,text="Execute",command=POSFileCheck,bg='burlywood1',width="10",font=("Tahoma", 12, 'bold'),justify='center')
#c7 = Button(master,text="Execute",bg='aqua',width="10",font=("Tahoma", 12, 'bold'),justify='center')
c7.place( height=40, width=200)
c7.grid(row = 12,column=2)

c1 = Button(master,text="Clear",command=ClearSourceFilePath,bg='light yellow',width="10",font=("Tahoma", 12, 'normal'),justify='center')
c2 = Button(master,text="Clear",command=ClearArchiveFilePath,bg='light goldenrod yellow',width="10",font=("Tahoma", 12, 'normal'),justify='center')
c3 = Button(master,text="Clear",command=ClearResultFilePath,bg='pale goldenrod',width="10",font=("Tahoma", 12, 'normal'),justify='center')
c4 = Button(master,text="Clear",command=ClearLogFolderPath,bg='khaki',width="10",font=("Tahoma", 12, 'normal'),justify='center')
c5 = Button(master,text="Clear",command=ClearFolderDate,bg='dark khaki',width="10",font=("Tahoma", 12, 'normal'),justify='center')
c6 = Button(master,text="Clear All",command=ClearAllFilePath,bg='ivory3',width="10",font=("Tahoma", 12, 'bold'),justify='center')
c1.grid(row=0, column = 3,sticky=E,padx=10, pady=5)
c2.grid(row=2, column = 3,sticky=E,padx=10, pady=5)
c3.grid(row=4, column = 3,sticky=E,padx=10, pady=5)
c4.grid(row=6, column = 3,sticky=E,padx=10, pady=5)
c5.grid(row=8, column = 3,sticky=E,padx=10, pady=5)
c6.grid(row=12, column = 1)
mainloop()

